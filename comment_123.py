
from logging import exception
import time
import pandas as pd
import datetime
import openpyxl
from openpyxl import load_workbook

try:
    start=time.time()
        
    df=pd.read_excel(r'D:\Salman\update.xlsx',header=1,na_filter=False)
    df.replace(to_replace=('  '),value='',inplace=True)

    i=0
    count_item=0
    max=2000

    now = datetime.datetime.now()
    todays_date=now.strftime("%Y-%m-%d %H:%M:%S")
    todays_date=pd.to_datetime(todays_date)

    for x in range(max):
        if df.loc[i,'Activity Status']=='On-Hold' and df.loc[i,'On-Hold / Cancelled due to SDU (Y/N)']=='':
                df.loc[i,'Comments']='Update reason on hold due to CU or SDU'
                count_item+=1
        
        if df.loc[i,'Activity Status']=='Cancelled' and df.loc[i,'On-Hold / Cancelled due to SDU (Y/N)']=='':
                df.loc[i,'Comments']='Update reason for Cancellation due to CU or SDU'
                count_item+=1
        
        if df.loc[i,'Activity Status']=='In-Progress' and df.loc[i,'Start Date']>todays_date:
                df.loc[i,'Comments']='Update start date as it is greater than today’s date'
                count_item+=1
            
        if df.loc[i,'Activity Status']=='In-Progress' and df.loc[i,'End Date']<todays_date:
                df.loc[i,'Comments']='Update PAT status as Successful /Unsuccessful'
                count_item+=1
            
        if df.loc[i,'Activity Status']==('Successful' or 'Unsuccessful') and df.loc[i,'Start Date']>todays_date:
            df.loc[i,'Comments']='Update Start date as needful'
            count_item+=1
        
        if df.loc[i,'Activity Status']==('Successful' or 'Unsuccessful') and df.loc[i,'End Date']>todays_date:
            df.loc[i,'Comments']='Update End date as needful'
            count_item+=1
        
        if df.loc[i,'Activity Status']=='Approved' and df.loc[i,'Start Date']<todays_date:
            df.loc[i,'Comments']='Update the PAT Status as needed or Change Start date as agreed with CU'
            count_item+=1
        
        if df.loc[i,'Activity Status']=='Approved' and (df.loc[i,'End Date']<todays_date and df.loc[i,'Comments']==''):
            df.loc[i,'Comments']='Update the PAT Status as needed or Change Start date as agreed with CU '#End Date
            count_item+=1
        
        if df.loc[i,'Activity Status']==('Review-LM' or 'Review-SME') and (df.loc[i,'Start Date'] and df.loc[i,'End Date'])<todays_date:
            df.loc[i,'Comments']='Update PAT dates as agreed with CU'
            count_item+=1
        
        if df.loc[i,'LNA Category']==('' or '0'):
                df.loc[i,'Comments']='Update LNA Category WFH Friendly or WFH Not Friendly'
                count_item+=1
        
        if df.loc[i,'LNA Category Comments']==('' or '0'):
                df.loc[i,'Comments']='Update LNA Category Comment'
                count_item+=1
        else:
            pass
        i+=1
        
    path=r'D:\Salman\update.xlsx'
    book = openpyxl.load_workbook(path)

    comment=0
    r=3
    for x in range(max):
        book.active.cell(row=r,column=51,value=df.Comments[comment])
        comment+=1
        r+=1
    book.save(r'D:\Salman\update.xlsx')

    print(count_item,' Comments updated successfully in Comments column!!')
    end=time.time()
    finish=end-start
    print('Process finished in '+str(finish)+' seconds')

except exception as e:
    print(e)

